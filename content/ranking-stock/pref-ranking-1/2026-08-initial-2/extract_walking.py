"""Extract the walking participation-rate ranking from the Statistics Bureau Excel."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
COMMON_DIR = BASE_DIR.parent / "common"
if str(COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_DIR))

from prefectures import PREFECTURE_BY_CODE, PREFECTURES  # noqa: E402


DEFAULT_OUTPUT_PATH = BASE_DIR / "data" / "walking.json"
SOURCE_NAME = (
    "総務省統計局「令和3年社会生活基本調査 47都道府県ランキング"
    "(ウォーキングが人気!?ランキング)」"
)


def competition_rank(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sort rows by value and assign competition ranks.

    Args:
        rows: Extracted prefecture rows with a numeric ``value`` field.

    Returns:
        Rows in descending-value order, with a ``rank`` field added.
    """
    ranked_rows: list[dict[str, Any]] = []
    previous_value: float | None = None
    previous_rank = 0
    for index, row in enumerate(
        sorted(rows, key=lambda item: (-item["value"], item["pref_code"])), start=1
    ):
        rank = previous_rank if row["value"] == previous_value else index
        ranked_rows.append({"rank": rank, **row})
        previous_value = row["value"]
        previous_rank = rank
    return ranked_rows


def require_unique_top_six(ranking: list[dict[str, Any]]) -> None:
    """Reject an item whose first six published values contain a tie."""
    top_six = ranking[:6]
    values = [entry["value"] for entry in top_six]
    if len(set(values)) != len(values):
        raise ValueError(f"TOP6 に同値があります: {top_six}")


def extract_rows(input_path: Path) -> list[dict[str, Any]]:
    """Read the 47 prefecture rows from the ``ランキング`` worksheet.

    Args:
        input_path: Downloaded Statistics Bureau Excel file.

    Returns:
        Prefecture code, name, value, and source rank for all 47 prefectures.
    """
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["ランキング"]
        rows: list[dict[str, Any]] = []
        nationwide: float | None = None
        pref_code_by_name = {pref.name: pref.code for pref in PREFECTURES}
        for row in worksheet.iter_rows(values_only=True):
            source_rank, pref_name, value = row[1:4]
            if pref_name == "全国平均":
                nationwide = float(value)
            elif pref_name in pref_code_by_name:
                if not isinstance(source_rank, int) or not isinstance(value, (int, float)):
                    raise ValueError(f"都道府県行の形式が不正です: {row}")
                rows.append(
                    {
                        "pref_code": pref_code_by_name[pref_name],
                        "pref_name": pref_name,
                        "value": float(value),
                        "source_rank": source_rank,
                    }
                )
    finally:
        workbook.close()

    if nationwide != 44.3:
        raise ValueError(f"全国平均が想定値 44.3 と一致しません: {nationwide}")
    if len(rows) != 47:
        raise ValueError(f"都道府県行は 47 件必要です: {len(rows)}")
    if {row["pref_code"] for row in rows} != set(range(1, 48)):
        raise ValueError("都道府県コードが 1〜47 をちょうど 1 回ずつ含んでいません")
    return rows


def build_data(input_path: Path) -> dict[str, Any]:
    """Build the canonical ranking JSON from a downloaded Excel file."""
    rows = extract_rows(input_path)
    ranking = competition_rank(rows)
    for entry in ranking:
        if entry["rank"] != entry.pop("source_rank"):
            raise ValueError(f"公表順位と計算順位が一致しません: {entry}")
    require_unique_top_six(ranking)

    return {
        "ranking_data": {
            "entries": [
                {
                    "rank": entry["rank"],
                    "pref_code": entry["pref_code"],
                    "value": entry["value"],
                }
                for entry in ranking[:5]
            ]
        },
        "source_note_text": (
            "都道府県値をそのまま使用（市区町村統計ではないため県換算なし）。"
            "値は 10 歳以上の行動者率（過去 1 年間に「ウォーキング・軽い体操」を"
            "した人の割合）。公表 Excel の 47 行を機械抽出し、TOP6 に同値がないことを"
            "検査済み。"
        ),
        "full_ranking": [
            {
                "rank": entry["rank"],
                "pref_code": entry["pref_code"],
                "pref_name": entry["pref_name"],
                "value": entry["value"],
            }
            for entry in ranking
        ],
        "meta": {
            "item": "ウォーキング・軽い体操の行動者率",
            "category": "スポーツ",
            "measure": "行動者率",
            "unit": "%",
            "data_year_label": "令和3年(2021年)",
            "source_name": SOURCE_NAME,
            "retrieved_on": "2026-08-24",
            "url": "https://www.stat.go.jp/data/shakai/2021/rank/zuhyou/rank/rank21.xlsx",
            "suffix": "%",
            "prefix": None,
            "nationwide": 44.3,
        },
    }


def main() -> None:
    """Parse CLI arguments and write the generated JSON."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_path", type=Path, help="入力の rank21.xlsx")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()

    data = build_data(args.input_path)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


if __name__ == "__main__":
    main()
