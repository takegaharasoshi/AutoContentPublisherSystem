"""Extract activity-rate rankings (basketball / gardening / manga) from Statistics Bureau Excels.

第 2 バッチの extract_walking.py と同じ出力スキーマで、第 3 バッチで使う
社会生活基本調査の 3 指標をまとめて抽出する。rank25 / rank27 は 1 シートに
2 つのランキング表が並ぶため、表ごとの列オフセットを ITEMS に持つ。
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
COMMON_DIR = BASE_DIR.parent / "common"
if str(COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_DIR))

from prefectures import PREFECTURES  # noqa: E402


RETRIEVED_ON = "2026-08-27"


ITEMS: dict[str, dict[str, Any]] = {
    "basketball": {
        "file": "rank28.xlsx",
        "rank_col": 1,
        "sheet_title": "バスケが人気！？ランキング",
        "item": "バスケットボールの行動者率",
        "definition": "過去 1 年間に「バスケットボール」をした人の割合",
        "nationwide": 3.6,
        "crosscheck": {1: ("秋田県", 5.3), 2: ("沖縄県", 5.1)},
        "url": "https://www.stat.go.jp/data/shakai/2021/rank/zuhyou/rank/rank28.xlsx",
    },
    "gardening": {
        "file": "rank25.xlsx",
        "rank_col": 1,
        "sheet_title": "園芸・ガーデニングが人気！？ランキング",
        "item": "園芸・庭いじり・ガーデニングの行動者率",
        "definition": "過去 1 年間に「園芸・庭いじり・ガーデニング」をした人の割合",
        "nationwide": 26.0,
        "crosscheck": {1: ("群馬県", 32.8), 2: ("長野県", 32.3)},
        "url": "https://www.stat.go.jp/data/shakai/2021/rank/zuhyou/rank/rank25.xlsx",
    },
    "manga": {
        "file": "rank27.xlsx",
        "rank_col": 6,
        "sheet_title": "マンガ大好き！？ランキング",
        "item": "マンガを読む人の割合",
        "definition": "過去 1 年間に「マンガを読む」に回答した人の割合",
        "nationwide": 36.8,
        "crosscheck": {1: ("東京都", 43.2), 2: ("神奈川県", 41.2)},
        "url": "https://www.stat.go.jp/data/shakai/2021/rank/zuhyou/rank/rank27.xlsx",
    },
}


def competition_rank(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sort rows by value and assign competition ranks.

    Args:
        rows: Extracted prefecture rows with a numeric ``value`` field.

    Returns:
        Rows in descending-value order, with a ``rank`` field added.
    """
    ranked_rows: list[dict[str, Any]] = []
    previous_value: float | None = None
    previous_rank = 0
    for index, row in enumerate(
        sorted(rows, key=lambda item: (-item["value"], item["pref_code"])), start=1
    ):
        rank = previous_rank if row["value"] == previous_value else index
        ranked_rows.append({"rank": rank, **row})
        previous_value = row["value"]
        previous_rank = rank
    return ranked_rows


def require_unique_top_six(ranking: list[dict[str, Any]]) -> None:
    """Reject an item whose first six published values contain a tie."""
    top_six = ranking[:6]
    values = [entry["value"] for entry in top_six]
    if len(set(values)) != len(values):
        raise ValueError(f"TOP6 に同値があります: {top_six}")


def extract_rows(
    input_path: Path, rank_col: int
) -> tuple[list[dict[str, Any]], float | None]:
    """Read the 47 prefecture rows of one ranking table on the worksheet.

    Args:
        input_path: Downloaded Statistics Bureau Excel file.
        rank_col: Zero-based column index of the table's 順位 column.

    Returns:
        Prefecture rows (code, name, value, source rank) and the 全国平均 value.
    """
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["ランキング"]
        rows: list[dict[str, Any]] = []
        nationwide: float | None = None
        pref_code_by_name = {pref.name: pref.code for pref in PREFECTURES}
        for row in worksheet.iter_rows(values_only=True):
            cells = list(row) + [None] * (rank_col + 3 - len(row))
            source_rank, pref_name, value = cells[rank_col : rank_col + 3]
            if pref_name == "全国平均" and isinstance(value, (int, float)):
                nationwide = float(value)
            elif isinstance(pref_name, str) and pref_name.strip() in pref_code_by_name:
                if not isinstance(source_rank, int) or not isinstance(value, (int, float)):
                    raise ValueError(f"都道府県行の形式が不正です: {row}")
                name = pref_name.strip()
                rows.append(
                    {
                        "pref_code": pref_code_by_name[name],
                        "pref_name": name,
                        "value": float(value),
                        "source_rank": source_rank,
                    }
                )
    finally:
        workbook.close()

    if len(rows) != 47:
        raise ValueError(f"都道府県行は 47 件必要です: {len(rows)}")
    if {row["pref_code"] for row in rows} != set(range(1, 48)):
        raise ValueError("都道府県コードが 1〜47 をちょうど 1 回ずつ含んでいません")
    return rows, nationwide


def build_data(name: str, input_dir: Path) -> dict[str, Any]:
    """Build the canonical ranking JSON for one configured item."""
    spec = ITEMS[name]
    input_path = input_dir / spec["file"]
    rows, nationwide = extract_rows(input_path, spec["rank_col"])
    if nationwide != spec["nationwide"]:
        raise ValueError(
            f"{name}: 全国平均が想定値 {spec['nationwide']} と一致しません: {nationwide}"
        )
    ranking = competition_rank(rows)
    for entry in ranking:
        if entry["rank"] != entry.pop("source_rank"):
            raise ValueError(f"{name}: 公表順位と計算順位が一致しません: {entry}")
    require_unique_top_six(ranking)
    for rank, (pref_name, value) in spec["crosscheck"].items():
        entry = ranking[rank - 1]
        if (entry["pref_name"], entry["value"]) != (pref_name, value):
            raise ValueError(
                f"{name}: クロスチェック不一致 rank={rank}: {entry} != {(pref_name, value)}"
            )

    return {
        "ranking_data": {
            "entries": [
                {
                    "rank": entry["rank"],
                    "pref_code": entry["pref_code"],
                    "value": entry["value"],
                }
                for entry in ranking[:5]
            ]
        },
        "source_note_text": (
            "都道府県値をそのまま使用（市区町村統計ではないため県換算なし）。"
            f"値は 10 歳以上の行動者率（{spec['definition']}）。"
            "公表 Excel の 47 行を機械抽出し、公表順位との一致・TOP6 に同値が"
            "ないことを検査済み。"
        ),
        "full_ranking": [
            {
                "rank": entry["rank"],
                "pref_code": entry["pref_code"],
                "pref_name": entry["pref_name"],
                "value": entry["value"],
            }
            for entry in ranking
        ],
        "meta": {
            "item": spec["item"],
            "category": "趣味・スポーツ",
            "measure": "行動者率",
            "unit": "%",
            "data_year_label": "令和3年(2021年)",
            "source_name": (
                "総務省統計局「令和3年社会生活基本調査 47都道府県ランキング"
                f"({spec['sheet_title']})」"
            ),
            "retrieved_on": RETRIEVED_ON,
            "url": spec["url"],
            "suffix": "%",
            "prefix": None,
            "nationwide": spec["nationwide"],
        },
    }


def main() -> None:
    """Parse CLI arguments and write the generated JSON files."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "input_dir", type=Path, help="rank25/27/28.xlsx を置いたディレクトリ"
    )
    parser.add_argument("--output-dir", type=Path, default=BASE_DIR / "data")
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    for name in ITEMS:
        data = build_data(name, args.input_dir)
        output_path = args.output_dir / f"{name}.json"
        output_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        top = data["full_ranking"][0]
        print(f"{name}: 1位 {top['pref_name']} {top['value']}% -> {output_path}")


if __name__ == "__main__":
    main()
