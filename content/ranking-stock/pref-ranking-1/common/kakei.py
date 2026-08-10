"""Download and parse Statistics Bureau Household Survey ranking workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Mapping
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
import unicodedata

from openpyxl import load_workbook

from prefectures import CITY_TO_PREF_CODE


RANK_URL_TEMPLATE = "https://www.stat.go.jp/data/kakei/zuhyou/rank{rank_no:02d}.xlsx"
USER_AGENT = "AutoContentPublisherSystem-pref-ranking-tool/1.0"
FIRST_DATA_ROW = 5
LAST_DATA_ROW = 57


@dataclass(frozen=True)
class WorkbookData:
    """Parsed data from one ranking workbook.

    Attributes:
        data_year_label: Display label read from cell A1.
        category: Workbook sheet name (major category).
        items: Per-item, per-measure series keyed by (item name, measure).
    """

    data_year_label: str
    category: str
    items: dict[tuple[str, str], ItemSeries]


@dataclass(frozen=True)
class ItemSeries:
    """One item and measurement series from a ranking workbook.

    Attributes:
        name: Published item name.
        measure: Measurement series name, either ``金額`` or ``数量``.
        unit: Unit used by the series, such as ``円``, ``kg``, or ``g``.
        values: Annual values keyed by the 52 published city names.
        nationwide: Nationwide value from rank 0.
    """

    name: str
    measure: str
    unit: str
    values: dict[str, int | Decimal]
    nationwide: int | Decimal


def download(rank_no: int, *, cache_dir: Path, force: bool = False) -> Path:
    """Download a ranking workbook into the local cache.

    Args:
        rank_no: Ranking workbook number, from 1 through 14.
        cache_dir: Directory that holds cached workbooks.
        force: Download again even when a cached file exists.

    Returns:
        Path to the cached workbook.

    Raises:
        ValueError: If rank_no is outside the available range.
        RuntimeError: If the download cannot be completed.
    """
    if rank_no not in range(1, 15):
        raise ValueError("rank は 1〜14 を指定してください")
    cache_dir.mkdir(parents=True, exist_ok=True)
    destination = cache_dir / f"rank{rank_no:02d}.xlsx"
    if destination.exists() and not force:
        return destination

    request = Request(
        RANK_URL_TEMPLATE.format(rank_no=rank_no),
        headers={"User-Agent": USER_AGENT},
    )
    temporary = destination.with_suffix(".xlsx.part")
    try:
        with urlopen(request, timeout=30) as response:
            temporary.write_bytes(response.read())
        temporary.replace(destination)
    except (HTTPError, URLError, OSError) as exc:
        temporary.unlink(missing_ok=True)
        raise RuntimeError(f"rank{rank_no:02d}.xlsx の取得に失敗しました: {exc}") from exc
    return destination


def parse(path: Path) -> WorkbookData:
    """Parse one Household Survey ranking workbook.

    Args:
        path: Path to an xlsx file in the published ranking-table layout.

    Returns:
        Parsed workbook data.

    Raises:
        ValueError: If the workbook layout or city data is invalid.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) != 1:
            raise ValueError("家計調査 Excel は 1 シートである必要があります")
        # read_only シートの cell() はランダムアクセスのたびに全走査するため、
        # 一度だけ iter_rows で読み切ってグリッドとして参照する。
        grid = _Grid(list(workbook.active.iter_rows(values_only=True)))
        data_year_label = str(grid.cell(1, 1) or "").strip()
        if not data_year_label:
            raise ValueError("A1 のデータ年ラベルが見つかりません")

        items: dict[tuple[str, str], ItemSeries] = {}
        for city_column in range(2, grid.max_column + 1, 2):
            item_name = grid.cell(3, city_column)
            if item_name is None or not str(item_name).strip():
                continue
            name = str(item_name).strip()
            measure, unit = _parse_measure_header(grid.cell(4, city_column), name)
            key = (name, measure)
            if key in items:
                raise ValueError(f"品目・系列が重複しています: {name}（{measure}）")
            city_values, national_value = _parse_item_column(
                grid, city_column, f"{name}（{measure}）", measure
            )
            items[key] = ItemSeries(
                name=name,
                measure=measure,
                unit=unit,
                values=city_values,
                nationwide=national_value,
            )
        if not items:
            raise ValueError("3 行目に品目名が見つかりません")
        return WorkbookData(data_year_label, workbook.sheetnames[0], items)
    finally:
        workbook.close()


def find_item(
    items: Mapping[tuple[str, str], ItemSeries],
    query: str,
    measure: str | None = None,
) -> ItemSeries:
    """Find one item name by normalized partial match.

    Args:
        items: Mapping of item and measurement series.
        query: Item text entered by a user.
        measure: Optional measurement series to select.

    Returns:
        The matching item series.

    Raises:
        ValueError: If there is no match or more than one match.
    """
    _validate_measure(measure)
    normalized_query = _normalize_for_search(query)
    if not normalized_query:
        raise ValueError("品目名を指定してください")
    matches = [
        series
        for series in items.values()
        if normalized_query in _normalize_for_search(series.name)
        and (measure is None or series.measure == measure)
    ]
    # 「パン」「外食」のように別品目の部分文字列になる名前があるため、
    # 完全一致が 1 件だけならそれを優先する。
    exact = [
        series
        for series in matches
        if _normalize_for_search(series.name) == normalized_query
    ]
    if len(exact) == 1:
        return exact[0]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise ValueError(f"品目が見つかりません: {query}")
    raise ValueError(f"品目が複数見つかりました: {query}（候補: {_series_labels(matches)}）")


def _parse_measure_header(header: object, item_name: str) -> tuple[str, str]:
    """Determine the measurement series and unit from a row-four header."""
    normalized_header = _normalize_for_search(str(header or ""))
    if "金額" in normalized_header:
        return "金額", "円"
    if "数量:" in normalized_header:
        unit_text = normalized_header.split("数量:", maxsplit=1)[1]
        unit = _normalize_for_search(unit_text.split(">", maxsplit=1)[0])
        if unit:
            return "数量", unit
    raise ValueError(f"{item_name}: 4 行目の系列ヘッダが不明です: {header}")


def _validate_measure(measure: str | None) -> None:
    """Validate an optional supported measurement filter."""
    if measure not in (None, "金額", "数量"):
        raise ValueError(f"未知の系列です: {measure}")


def _series_labels(series_list: list[ItemSeries]) -> str:
    """Build a compact candidate list for an ambiguous item search."""
    return ", ".join(
        f"{series.name}（{series.measure}・{series.unit}）"
        for series in series_list
    )


class _Grid:
    """Random-access view over the cell values of a worksheet read in one pass."""

    def __init__(self, rows: list[tuple[object, ...]]) -> None:
        self._rows = rows
        self.max_column = max((len(row) for row in rows), default=0)

    def cell(self, row: int, column: int) -> object:
        """Return the value at a 1-indexed position, or None when out of range."""
        if not 1 <= row <= len(self._rows):
            return None
        values = self._rows[row - 1]
        if not 1 <= column <= len(values):
            return None
        return values[column - 1]


def _parse_item_column(
    grid: _Grid, city_column: int, item_name: str, measure: str
) -> tuple[dict[str, int | Decimal], int | Decimal]:
    """Parse and validate the 53 ranking rows for a single item."""
    city_values: dict[str, int | Decimal] = {}
    national_value: int | Decimal | None = None
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        rank = grid.cell(row, 1)
        city = grid.cell(row, city_column)
        amount = _as_value(grid.cell(row, city_column + 1), item_name, row, measure)
        if rank == 0:
            if national_value is not None:
                raise ValueError(f"{item_name}: 全国値が重複しています")
            national_value = amount
            continue
        if not isinstance(city, str) or not city.strip():
            raise ValueError(f"{item_name}: {row} 行目の市名が不正です")
        city_name = city.strip()
        if city_name not in CITY_TO_PREF_CODE:
            raise ValueError(f"{item_name}: 未知の市名です: {city_name}")
        if city_name in city_values:
            raise ValueError(f"{item_name}: 市名が重複しています: {city_name}")
        city_values[city_name] = amount

    missing = set(CITY_TO_PREF_CODE) - set(city_values)
    unexpected = set(city_values) - set(CITY_TO_PREF_CODE)
    if national_value is None or len(city_values) != 52 or missing or unexpected:
        details = []
        if national_value is None:
            details.append("全国値なし")
        if missing:
            details.append(f"不足: {', '.join(sorted(missing))}")
        if unexpected:
            details.append(f"未知: {', '.join(sorted(unexpected))}")
        raise ValueError(f"{item_name}: 52 市のデータが揃っていません（{' / '.join(details)}）")
    return city_values, national_value


def _as_value(
    value: object, item_name: str, row: int, measure: str
) -> int | Decimal:
    """Convert an Excel value while retaining decimal quantities."""
    if isinstance(value, bool):
        raise ValueError(f"{item_name}: {row} 行目の値が不正です")
    try:
        numeric = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"{item_name}: {row} 行目の値が不正です") from exc
    if not numeric.is_finite():
        raise ValueError(f"{item_name}: {row} 行目の値が不正です")
    if measure == "数量":
        return numeric if numeric != numeric.to_integral_value() else int(numeric)
    integer = int(numeric)
    if numeric != integer:
        raise ValueError(f"{item_name}: {row} 行目の金額が整数ではありません")
    return integer


def _normalize_for_search(value: str) -> str:
    """Normalize width variants and remove whitespace for item matching."""
    return "".join(unicodedata.normalize("NFKC", value).split())
