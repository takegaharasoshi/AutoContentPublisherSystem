"""Tests for Household Survey workbook download and parsing."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import kakei  # noqa: E402
from prefectures import CITY_TO_PREF_CODE  # noqa: E402


def make_ranking_workbook(path: Path, item_name: str = "ｱｲｽｸﾘｰﾑ･ｼｬｰﾍﾞｯﾄ") -> None:
    """Create a minimal xlsx that follows the published ranking layout."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "菓子類"
    worksheet["A1"] = "2023年（令和５年）～2025年（令和７年）平均"
    worksheet["B3"] = item_name
    worksheet["B4"] = "----------<金 額>-"
    worksheet["A5"] = 0
    worksheet["B5"] = "全国"
    worksheet["C5"] = 999
    for rank, city in enumerate(CITY_TO_PREF_CODE, start=1):
        row = rank + 5
        worksheet.cell(row, 1, rank)
        worksheet.cell(row, 2, city)
        worksheet.cell(row, 3, 100 + rank)
    workbook.save(path)
    workbook.close()


def make_mixed_measure_workbook(path: Path) -> None:
    """Create an xlsx with same-name amount and quantity series."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "穀類"
    worksheet["A1"] = "2023年（令和５年）～2025年（令和７年）平均"
    series = (
        (2, "米", "----------<金　額>-", 1000),
        (4, "米", "--------<数量：ｋｇ>-", 10.5),
        (6, "パン", "--------<数量：　ｇ>-", 100),
    )
    for column, name, header, nationwide in series:
        worksheet.cell(3, column, name)
        worksheet.cell(4, column, header)
        worksheet.cell(5, column, "全国")
        worksheet.cell(5, column + 1, nationwide)
        for rank, city in enumerate(CITY_TO_PREF_CODE, start=1):
            row = rank + 5
            worksheet.cell(row, 1, rank)
            worksheet.cell(row, column, city)
            worksheet.cell(row, column + 1, nationwide + rank)
    worksheet["A5"] = 0
    workbook.save(path)
    workbook.close()


def test_parse_reads_year_category_and_city_values(tmp_path: Path) -> None:
    """Parser separates the nationwide row and retains all city values."""
    path = tmp_path / "rank09.xlsx"
    make_ranking_workbook(path)

    data = kakei.parse(path)

    assert data.data_year_label == "2023年（令和５年）～2025年（令和７年）平均"
    assert data.category == "菓子類"
    series = data.items[("ｱｲｽｸﾘｰﾑ･ｼｬｰﾍﾞｯﾄ", "金額")]
    assert series.nationwide == 999
    assert len(series.values) == 52
    assert series.values["札幌市"] == 101


def test_find_item_normalizes_width_and_whitespace() -> None:
    """Full-width query text matches half-width source item text."""
    items = {
        ("ｱｲｽｸﾘｰﾑ･ｼｬｰﾍﾞｯﾄ", "金額"): kakei.ItemSeries(
            "ｱｲｽｸﾘｰﾑ･ｼｬｰﾍﾞｯﾄ", "金額", "円", {}, 0
        ),
        ("チョコ レート", "金額"): kakei.ItemSeries(
            "チョコ レート", "金額", "円", {}, 0
        ),
    }

    assert (
        kakei.find_item(items, "アイスクリーム ・ シャーベット").name
        == "ｱｲｽｸﾘｰﾑ･ｼｬｰﾍﾞｯﾄ"
    )
    with pytest.raises(ValueError, match="候補"):
        kakei.find_item(
            {
                ("菓子類", "金額"): kakei.ItemSeries("菓子類", "金額", "円", {}, 0),
                ("菓子パン", "金額"): kakei.ItemSeries("菓子パン", "金額", "円", {}, 0),
            },
            "菓子",
        )


def test_parse_separates_measure_series_and_normalizes_units(tmp_path: Path) -> None:
    """Amount and quantity series with the same name are kept separately."""
    path = tmp_path / "rank01.xlsx"
    make_mixed_measure_workbook(path)

    data = kakei.parse(path)

    assert data.items[("米", "金額")].unit == "円"
    assert data.items[("米", "数量")].unit == "kg"
    assert data.items[("パン", "数量")].unit == "g"
    assert data.items[("米", "金額")].values["札幌市"] == 1001
    assert data.items[("米", "数量")].values["札幌市"] == Decimal("11.5")
    assert kakei.find_item(data.items, "米", "数量").unit == "kg"
    with pytest.raises(ValueError, match="数量・kg"):
        kakei.find_item(data.items, "米")


def test_parse_rejects_unknown_city(tmp_path: Path) -> None:
    """Parser reports a city not in the specified 52-city mapping."""
    path = tmp_path / "invalid.xlsx"
    make_ranking_workbook(path)
    workbook = load_workbook(path)
    workbook.active["B6"] = "架空市"
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="未知の市名"):
        kakei.parse(path)


def test_parse_rejects_unknown_measure_header(tmp_path: Path) -> None:
    """An unrecognized row-four header is never silently imported."""
    path = tmp_path / "invalid-header.xlsx"
    make_ranking_workbook(path)
    workbook = load_workbook(path)
    workbook.active["B4"] = "----------<指数>-"
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="系列ヘッダが不明"):
        kakei.parse(path)


def test_download_uses_cache_and_user_agent(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Download writes xlsx bytes once and then reuses its cache."""
    requests: list[object] = []

    class Response:
        def __enter__(self) -> "Response":
            return self

        def __exit__(self, *args: object) -> None:
            return None

        def read(self) -> bytes:
            return b"xlsx-bytes"

    def fake_urlopen(request: object, timeout: int) -> Response:
        requests.append((request, timeout))
        return Response()

    monkeypatch.setattr(kakei, "urlopen", fake_urlopen)
    first_path = kakei.download(9, cache_dir=tmp_path)
    second_path = kakei.download(9, cache_dir=tmp_path)

    assert first_path == second_path
    assert first_path.read_bytes() == b"xlsx-bytes"
    assert len(requests) == 1
    request, timeout = requests[0]
    assert request.get_header("User-agent") == kakei.USER_AGENT
    assert timeout == 30
